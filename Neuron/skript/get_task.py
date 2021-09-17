import openpyxl
from langdetect import detect

def get_task():
	wb = openpyxl.load_workbook("D:/Работы ТГУ/Научка/Adaptive/Neuron/Tables/TaskTranslation.xlsx")

	sheet = wb["TaskTranslation_ANSI"]

	rows = sheet.max_row


	mas_id = []
	mas_name = []

	for i_row in range(1,rows+1):
		cell = sheet.cell(row = i_row, column=openpyxl.utils.column_index_from_string("A"))
		mas_id.append(cell.value)

	for i_row in range(1,rows+1):
		cell = sheet.cell(row = i_row, column=openpyxl.utils.column_index_from_string("B"))
		mas_name.append(cell.value)

	mas_end = []

	for i in range(0,len(mas_id)):
		if mas_id[i] != "-":
			tm = mas_id[i]
			mas_id[i] == "-"
			mas = [tm,mas_name[i]]
			for g in range(0,len(mas_id)):
				if i != g and tm == mas_id[g]:
					mas.append(mas_name[g])

			mas_end.append(mas)

	mas_end_2 = []
	for i in range(0,len(mas_end)):
		ez1 = detect(mas_end[i][1])
		ez2 = ""
		if len(mas_end[i]) != 2:
			ez2 = detect(mas_end[i][2])

		mas = []
		if (ez1 != 'ru' and ez2 != 'en') and len(mas_end[i]) != 2:
			mas.append(mas_end[i][0])
			mas.append(mas_end[i][2])
			mas.append(mas_end[i][1])
		elif len(mas_end[i]) != 2:
			mas.append(mas_end[i][0])
			mas.append(mas_end[i][1])
			mas.append(mas_end[i][2])
		else:
			mas.append(mas_end[i][0])
			mas.append(mas_end[i][1])

		mas_end_2.append(mas)


	print(mas_end_2)
	wb = openpyxl.Workbook()
	wb.create_sheet(title = 'Первый лист', index = 0)
	sheet = wb['Первый лист']

	row_counter = 1
	for i in mas_end_2:
		for col in range(1,len(i)+1):
			cell = sheet.cell(row = row_counter, column=col)
			cell.value = i[col-1]

		row_counter+= 1

	wb.save('example.xlsx')
		
get_task()


