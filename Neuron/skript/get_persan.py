import openpyxl

def get_person():
	wb = openpyxl.load_workbook("D:/Работы ТГУ/Научка/Adaptive/Neuron/Tables/Result.xlsx")

	sheet = wb["Result_2_ANSI"]

	rows = sheet.max_row

	mas_name = []
	mas_task = []

	for i_row in range(1,rows+1):
		cell = sheet.cell(row = i_row, column=openpyxl.utils.column_index_from_string("B"))
		mas_name.append(cell.value)

	for i_row in range(1,rows+1):
		cell = sheet.cell(row = i_row, column=openpyxl.utils.column_index_from_string("C"))
		mas_task.append(cell.value)

	mas_end = []

	for i in range(0,len(mas_name)):
		if mas_name[i] != "-":
			tm = mas_name[i]
			mas_name[i] = "-"
			mas = [tm,mas_task[i]]
			for g in range(0,len(mas_name)):
				if i != g and tm == mas_name[g]:
					mas_name[g] = "-"
					flag = True
					for k in mas:
						if k == mas_task[g]:
							flag = False
					if flag: 
						mas.append(mas_task[g])

			mas_end.append(mas)


	wb = openpyxl.Workbook()
	wb.create_sheet(title = 'Первый лист', index = 0)
	sheet = wb['Первый лист']

	col_counter = 1
	for i in mas_end:
		for row in range(1,len(i)+1):
			cell = sheet.cell(row = row, column=col_counter)
			cell.value = i[row-1]

		col_counter+= 1

	wb.save('example.xlsx')

get_person()