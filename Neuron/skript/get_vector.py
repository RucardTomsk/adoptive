import openpyxl

def get_vector():
	wb = openpyxl.load_workbook("D:/Работы ТГУ/Научка/Adaptive/Neuron/Tables/ResultEnd.xlsx")
	wb1 = openpyxl.load_workbook("D:/Работы ТГУ/Научка/Adaptive/Neuron/Tables/TestTasksEnd.xlsx")

	sheet = wb["Первый лист"]
	sheet1 = wb1["TestTasks"]

	col = sheet.max_column
	max_row_task = sheet1.max_row
	row = sheet.max_row

	mas_name_task = []
	for i_col in range(1,col+1):
		mas = []
		for i_row in range(1,row+1):
			cell = sheet.cell(row=i_row, column=i_col)
			mas.append(cell.value)
		mas_name_task.append(mas)

	mas_id = []
	mas_number = []
	for i_row in range(2, max_row_task+1):
		cell_id = sheet1.cell(row = i_row, column=openpyxl.utils.column_index_from_string("A"))
		cell_number = sheet1.cell(row = i_row, column=openpyxl.utils.column_index_from_string("C"))

		mas_id.append(cell_id.value)
		mas_number.append(cell_number.value)

	mas_end = []
	print(mas_id)
	print(mas_number)

	for mas_persen in mas_name_task:
		vector = [0]*257
		mas = [mas_persen[0]]
		for i in range(1,len(mas_persen)):
			task_number = -1
			for g in range(len(mas_id)):
				if mas_persen[i] == mas_id[g] and mas_id[g] != None:
					task_number = mas_number[g]

			if task_number != -1:
				vector[int(task_number)-1] = 1

		if(sum(vector) < 20):
			continue
		str_vector = ""
		for i in range(0,124):
			str_vector+= str(vector[i])

		mas.append(str_vector)
		mas_end.append(mas)



	wb = openpyxl.Workbook()
	wb.create_sheet(title = 'Первый лист', index = 0)
	sheet = wb['Первый лист']

	row_counter = 1
	for i in mas_end:
		for col in range(1,len(i)+1):
			cell = sheet.cell(row = row_counter, column=col)
			cell.value = i[col-1]

		row_counter+= 1

	wb.save('vector_1model.xlsx')	

get_vector()


